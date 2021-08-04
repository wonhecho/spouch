
from openpyxl import load_workbook
import calendar
import datetime
from datetime import timedelta
import random



print(calendar.calendar(2020))
month=int(input("개막하는 달을 입력해주세요"))
date = int(input("개막하는 일을 입력해주세요"))
start=datetime.date(2020,month,date)
d=0


load_wb = load_workbook("팀.xlsx",data_only=True)

load_ws = load_wb['Sheet1']

print(load_ws['A1'].value)

print(load_ws.cell(1,2).value)


all_values=[]
for row in load_ws.columns:
     row_value=[]
     for cell in row:
          row_value.append(cell.value)

     all_values.append(row_value)



str1 = '정왕일요2부'
str2 = '정왕일요3부'
str3 = '정왕토요3부'
str4 = '정왕토요4부'
Teamnum = len(all_values[0])
leasun2 = []
leasun3 = []
leasat3 = []
leasat4 = []
leasun=[]

count=0
month = [1,2,3,4,5,6,7,8,9,10,11,12]


for i in all_values[2]:
     if not i:
          continue
     if str1 in i:
          leasun2.append(all_values[1][count])
     if str2 in i:
          leasun3.append(all_values[1][count])
     if str3 in i:
          leasat3.append(all_values[1][count])
     if str4 in i:
          leasat4.append(all_values[1][count])

     count=count+1
print(str1,": ",leasun2)
print(str2,": ",leasun3)
print(str3,": ",leasat3)
print(str4,": ",leasat4)

sun2 = len(leasun2)
sun3 = len(leasun3)
sat3 = len(leasat3)
sat4 = len(leasat4)
sun2m = sun2-1
sun2h = (int)(sun2/2)
sun3m = sun3-1
sun3h = (int)(sun3/2)
sat3m = sat3-1
sat3h = (int)(sat3/2)
sat4m = sat4-1
sat4h = (int)(sat4/2)

number=int(input("(1. 정왕 일요2,3부 3.정왕 토요 3부 4. 정왕 일요 4부) 보고싶은 리그의 숫자를 입력하세요 :  "))
matches = int(input("경기 수를 입력하세요: "))
random.shuffle(leasun2)
random.shuffle(leasun3)

if number==1:

     team_id=[]
     for i in range(0,sun2):
          team_id.append(i)
     for i in range(0, matches):
          for j in range(0,sun2h):
               # print(leasun2[team_id[j]],":",leasun2[team_id[sun2-j-1]])
               if i%2==1:
                    leasun.append(leasun2[team_id[sun2-j-1]]+" : "+leasun2[team_id[j]])
               else:     
                    leasun.append(leasun2[team_id[j]]+" : "+leasun2[team_id[sun2-j-1]])
          for j in range(0,sun2m):
               team_id[j]=(team_id[j]+1)%(sun2m)
     print("\n") 
     team_id = []
     for i in range(0, sun3):
          team_id.append(i)
     for i in range(0, matches):

          for j in range(0, sun3h):
               # print(leasun3[team_id[j]], ":", leasun3[team_id[sun3- j-1]])
               if i%2==1:
                    leasun.append(leasun3[team_id[sun3-j-1]]+" : "+leasun3[team_id[j]])
               else:     
                    leasun.append(leasun3[team_id[j]]+" : "+leasun3[team_id[sun3-j-1]])
               
          for j in range(0, sun3m):
               team_id[j] = (team_id[j] + 1) % (sun3m)
         
     print("\n") 



elif number==3:
     team_id = []

     for i in range(0, sat3):
          team_id.append(i)
     for i in range(0, matches):
          print(i + 1, "일차")
          for j in range(0, sat3h):
               print(leasat3[team_id[j]], ":", leasat3[team_id[sat3 - j - 1]])

          for j in range(0, sat3m):
               team_id[j] = (team_id[j] + 1) % (sat3m)
     print("\n")

elif number==4:
     team_id = []

     for i in range(0, sat4):
          team_id.append(i)
     for i in range(0, matches):
          print(i + 1, "일차")\

          for j in range(0, sat4h):
               print(leasat4[team_id[j]], ":", leasat4[team_id[sat4 - j - 1]])

          for j in range(0, sat4m):
               team_id[j] = (team_id[j] + 1) % (sat4m)
     print("\n")
time = ["12시","2시","4시","6시"]

j=0
T = (int)(len(leasun)/2)
for i in range(0,T):
     if i==0:
          print(start)
    
     elif j==4:
          d=d+7
          td = timedelta(days=d)
          season = start+td
          if season.day==21:
               print("yes")
          
          print(season)
          j=0
       
     print(time[j]+" "+leasun[i])
     j=j+1
     print(time[j]+" "+leasun[i+T])
     j=j+1
   
